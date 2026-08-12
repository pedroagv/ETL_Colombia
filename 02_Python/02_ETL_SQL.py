import os
import sys
import glob
import io
import re
import csv
import math
import shutil
import zipfile
import logging
import argparse
import tempfile
import unicodedata
from datetime import datetime
from dotenv import load_dotenv

import pandas as pd
import pymysql
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Arquitectura de esta fase:
#   Python actúa únicamente como un puente de carga liviano ("dumb pipe").
#   Los archivos (Excel/CSV, incluso dentro de ZIP) se leen en bloques
#   (streaming) y cada bloque se inserta de inmediato en la tabla temporal
#   correspondiente, sin acumular el archivo completo ni listas gigantes de
#   tuplas en memoria. Las columnas de datos crudos se cargan como TEXT tal
#   cual vienen del archivo origen: la limpieza, conversión de tipos, filtros
#   y cruces se delegan por completo a SQL (sentencias/procedimientos
#   almacenados) en una fase posterior, no a este script.
# ---------------------------------------------------------------------------

# Cargar variables de entorno desde .env
load_dotenv()

# Configuración de Rutas y Variables
IMPO_DIR = os.path.expanduser(os.getenv("IMPO_DIR", "Descargas/Importacion"))
EXPO_DIR = os.path.expanduser(os.getenv("EXPO_DIR", "Descargas/Exportacion"))

# Configuración de MySQL
MYSQL_HOST = os.getenv("MYSQL_HOST", "localhost")
MYSQL_PORT = int(os.getenv("MYSQL_PORT", "3306"))
MYSQL_USER = os.getenv("MYSQL_USER", "root")
MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "")
MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "colombia")

LOG_LEVEL_STR = os.getenv("LOG_LEVEL", "INFO").upper()

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(line_buffering=True)

# Configurar Logging
logging_level = getattr(logging, LOG_LEVEL_STR, logging.INFO)
logging.basicConfig(
    level=logging_level,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("dian_etl_sql.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ],
    force=True
)


logger = logging.getLogger("dian_etl_sql")


def clean_column_name(column_name) -> str:
    """
    Limpia y normaliza el nombre de una columna para que sea válido en MySQL.
    - Quita acentos y caracteres diacríticos.
    - Convierte a minúsculas.
    - Reemplaza espacios y caracteres especiales por guiones bajos.
    """
    if not isinstance(column_name, str):
        column_name = str(column_name)

    # Normalizar acentos
    nfkd = unicodedata.normalize('NFKD', column_name)
    cleaned = "".join([c for c in nfkd if not unicodedata.combining(c)])

    # Convertir a minúsculas
    cleaned = cleaned.lower().strip()

    # Reemplazar espacios y caracteres no alfanuméricos por '_'
    cleaned = re.sub(r'[^a-z0-9_]', '_', cleaned)

    # Reemplazar guiones bajos consecutivos por uno solo
    cleaned = re.sub(r'_+', '_', cleaned)

    # Eliminar guiones bajos al inicio o final
    cleaned = cleaned.strip('_')

    # Si empieza por un número, anteponer 'col_'
    if cleaned and cleaned[0].isdigit():
        cleaned = f"col_{cleaned}"

    # Si la columna quedó vacía
    if not cleaned:
        cleaned = "campo_desconocido"

    return cleaned


def _clean_value(x):
    """Normaliza NaN/inf a None; deja pasar el resto de valores tal cual (crudos)."""
    if x is None:
        return None
    if isinstance(x, float) and (math.isnan(x) or math.isinf(x)):
        return None
    return x


class MySQLManager:
    def __init__(self):
        self.host = MYSQL_HOST
        self.port = MYSQL_PORT
        self.user = MYSQL_USER
        self.password = MYSQL_PASSWORD
        self.db_name = MYSQL_DATABASE
        self._ensure_database_exists()


    def get_connection(self, select_db=True):
        return pymysql.connect(
            host=self.host,
            port=self.port,
            user=self.user,
            password=self.password,
            database=self.db_name if select_db else None,
            charset="utf8mb4",
            autocommit=False
        )

    def _ensure_database_exists(self):
        """
        Crea la base de datos si no existe en el servidor MySQL local.
        """
        try:
            conn = self.get_connection(select_db=False)
            with conn.cursor() as cursor:
                sql = f"CREATE DATABASE IF NOT EXISTS `{self.db_name}` DEFAULT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;"
                cursor.execute(sql)
            conn.commit()
            conn.close()
            logger.info(f"Base de datos MySQL '{self.db_name}' verificada/creada correctamente.")
        except Exception as e:
            logger.error(f"Error al verificar/crear la base de datos '{self.db_name}': {e}")
            raise e

    def _rename_legacy_tables(self):
        """
        Renombra tablas de Colombia previamente existentes (e.g. importacion, exportacion,
        colombia_impo, colombia_expo, impo, expo) a 'temporal_impo' y 'temporal_expo'.
        """
        legacy_mappings = {
            "temporal_impo": ["importacion", "colombia_impo", "impo", "colombia_importacion"],
            "temporal_expo": ["exportacion", "colombia_expo", "expo", "colombia_exportacion"],
        }
        try:
            conn = self.get_connection()
            with conn.cursor() as cursor:
                cursor.execute("SHOW TABLES;")
                existing_tables = set(row[0] for row in cursor.fetchall())

                for target_table, legacy_names in legacy_mappings.items():
                    if target_table in existing_tables:
                        continue
                    for old_name in legacy_names:
                        if old_name in existing_tables:
                            logger.info(f"Renombrando tabla MySQL legacy '{old_name}' -> '{target_table}'...")
                            cursor.execute(f"RENAME TABLE `{old_name}` TO `{target_table}`;")
                            conn.commit()
                            existing_tables.remove(old_name)
                            existing_tables.add(target_table)
                            break
            conn.close()
        except Exception as e:
            logger.warning(f"Advertencia al verificar/renombrar tablas legacy en MySQL: {e}")

    def get_existing_columns(self, table_name: str) -> dict:
        """
        Retorna un diccionario con las columnas existentes en una tabla de MySQL: {col_name: col_type}.
        """
        conn = self.get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
                rows = cursor.fetchall()
                return {row[0]: row[1].lower() for row in rows}
        except pymysql.MySQLError:
            return {}
        finally:
            conn.close()

    def truncate_table(self, table_name: str):
        """
        Vacía la tabla temporal al iniciar el ciclo: una vez que las Fases 3-4
        terminan de consumir un archivo, ya quedó cargado en 'importacion' y
        encolado para Elasticsearch, así que no hace falta conservar la copia
        cruda y la temporal no debe crecer indefinidamente con cada corrida.
        No falla si la tabla todavía no existe (primera corrida).
        """
        conn = self.get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("SHOW TABLES LIKE %s", (table_name,))
                if not cursor.fetchone():
                    return
                cursor.execute(f"TRUNCATE TABLE `{table_name}`;")
            conn.commit()
            logger.info(f"Tabla temporal '{table_name}' truncada al iniciar el ciclo.")
        except Exception as e:
            conn.rollback()
            logger.error(f"Error truncando la tabla temporal '{table_name}': {e}")
            raise e
        finally:
            conn.close()

    def ensure_index(self, table_name: str, column: str = "archivo_origen"):
        """
        Verifica si existe un índice sobre la columna usada para la limpieza de idempotencia
        y lo crea si falta, sin afectar los datos existentes.
        """
        existing_cols = self.get_existing_columns(table_name)
        if column not in existing_cols:
            return

        conn = self.get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute(f"SHOW INDEX FROM `{table_name}` WHERE Column_name = %s", (column,))
                if cursor.fetchone():
                    return
                index_name = f"idx_{column}"
                logger.info(f"Creando índice '{index_name}' sobre '{table_name}.{column}' para acelerar las búsquedas por archivo...")
                cursor.execute(f"ALTER TABLE `{table_name}` ADD INDEX `{index_name}` (`{column}`(191));")
                conn.commit()
        except Exception as e:
            conn.rollback()
            logger.warning(f"No se pudo crear el índice sobre '{table_name}.{column}': {e}")
        finally:
            conn.close()

    def sync_table_schema(self, table_name: str, data_columns: list, cursor) -> str:
        """
        Crea la tabla en MySQL si no existe (todas las columnas de datos como TEXT,
        sin inferencia de tipos) o agrega las columnas nuevas que aparezcan.
        No hace ninguna limpieza/conversión: eso queda para SQL en una fase posterior.
        Retorna el nombre de la columna a usar para el borrado por idempotencia.
        """
        existing_cols = self.get_existing_columns(table_name)

        if not existing_cols:
            col_definitions = ["`id` BIGINT AUTO_INCREMENT PRIMARY KEY"]
            for col in data_columns:
                col_definitions.append(f"`{col}` TEXT")
            col_definitions += [
                "`anio_proceso` INT",
                "`mes_proceso` VARCHAR(20)",
                "`archivo_origen` VARCHAR(255)",
                "`fecha_carga` DATETIME",
                "INDEX `idx_archivo_origen` (`archivo_origen`(191))",
            ]
            create_sql = f"CREATE TABLE `{table_name}` (\n  " + ",\n  ".join(col_definitions) + "\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;"
            logger.info(f"Creando tabla '{table_name}' en MySQL (columnas crudas en TEXT)...")
            cursor.execute(create_sql)
            return "archivo_origen"

        for col in data_columns:
            if col not in existing_cols:
                logger.info(f"Agregando nueva columna cruda '{col}' (TEXT) a la tabla '{table_name}'...")
                cursor.execute(f"ALTER TABLE `{table_name}` ADD COLUMN `{col}` TEXT;")
                existing_cols[col] = "text"

        if "archivo_origen" in existing_cols:
            return "archivo_origen"
        if "nombre_archivo" in existing_cols:
            return "nombre_archivo"
        return "archivo_origen"

    def load_file_chunks(self, table_name: str, chunks_iter, anio_proceso, mes_proceso, archivo_origen: str) -> int:
        """
        Recorre un iterador de bloques (columns, rows) y los inserta en MySQL de forma
        incremental: sincroniza el esquema con el primer bloque, borra registros previos
        del mismo archivo (idempotencia) y luego inserta cada bloque apenas llega, usando
        una única conexión/transacción por bloque. Nunca acumula datos de bloques anteriores.
        """
        total_inserted = 0
        fecha_carga = datetime.now()
        meta = (anio_proceso, mes_proceso, archivo_origen, fecha_carga)

        conn = self.get_connection()
        try:
            with conn.cursor() as cursor:
                insert_sql = None

                for columns, rows in chunks_iter:
                    if not rows:
                        continue

                    if insert_sql is None:
                        delete_col = self.sync_table_schema(table_name, columns, cursor)
                        conn.commit()

                        deleted = cursor.execute(
                            f"DELETE FROM `{table_name}` WHERE `{delete_col}` = %s;", (archivo_origen,)
                        )
                        conn.commit()
                        if deleted:
                            logger.info(f"Limpieza previa: eliminados {deleted} registros previos de '{archivo_origen}' en '{table_name}'.")

                        all_cols = columns + ["anio_proceso", "mes_proceso", "archivo_origen", "fecha_carga"]
                        quoted_cols = ", ".join(f"`{c}`" for c in all_cols)
                        placeholders = ", ".join(["%s"] * len(all_cols))
                        insert_sql = f"INSERT INTO `{table_name}` ({quoted_cols}) VALUES ({placeholders})"

                    batch = [tuple(_clean_value(v) for v in row) + meta for row in rows]
                    cursor.executemany(insert_sql, batch)
                    conn.commit()
                    total_inserted += len(batch)
                    if total_inserted % 25000 == 0 or total_inserted == len(batch):
                        logger.info(f"Avance staging '{archivo_origen}': {total_inserted} filas insertadas en '{table_name}'.")
                    else:
                        logger.debug(f"Insertados {total_inserted} registros en '{table_name}'.")


            return total_inserted
        except Exception as e:
            conn.rollback()
            logger.error(f"Error insertando datos en la tabla '{table_name}': {e}")
            raise e
        finally:
            conn.close()


def pick_valid_member(z: zipfile.ZipFile) -> str:
    """
    Elige el primer archivo procesable (.xlsx, .xls, .csv, .txt) dentro del ZIP,
    sin descomprimirlo todavía.
    """
    valid_files = [
        f for f in z.namelist()
        if f.lower().endswith(('.xlsx', '.xls', '.csv', '.txt')) and not f.startswith('__MACOSX')
    ]
    if not valid_files:
        raise ValueError("No se encontraron archivos válidos (.xlsx, .xls, .csv, .txt) dentro del ZIP")
    return valid_files[0]


def _detect_csv_dialect(sample_bytes: bytes):
    """
    Detecta encoding y delimitador a partir de una muestra pequeña de bytes,
    sin necesidad de cargar el archivo completo.
    """
    for encoding in ("utf-8-sig", "latin1"):
        try:
            sample_text = sample_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        sample_text = sample_bytes.decode("latin1", errors="replace")
        encoding = "latin1"

    try:
        delimiter = csv.Sniffer().sniff(sample_text, delimiters=";,|\t").delimiter
    except csv.Error:
        counts = {d: sample_text.count(d) for d in (";", ",", "|", "\t")}
        delimiter = max(counts, key=counts.get) if any(counts.values()) else ","

    return delimiter, encoding


def stream_csv_chunks(zip_path: str, member_name: str, chunk_size: int):
    """
    Lee un CSV/TXT directamente desde el ZIP en bloques (chunksize) usando el
    motor C de pandas, sin cargar el archivo completo en memoria ni bufferizar
    todo su contenido descomprimido.
    """
    with zipfile.ZipFile(zip_path, "r") as z:
        with z.open(member_name) as f:
            sample = f.read(8192)

    delimiter, encoding = _detect_csv_dialect(sample)
    logger.debug(f"CSV/TXT detectado: delimiter='{delimiter}' encoding='{encoding}'")

    with zipfile.ZipFile(zip_path, "r") as z:
        with z.open(member_name) as raw:
            text_stream = io.TextIOWrapper(raw, encoding=encoding, errors="replace")
            reader = pd.read_csv(
                text_stream,
                sep=delimiter,
                dtype=str,
                chunksize=chunk_size,
                on_bad_lines="skip",
            )
            for chunk_df in reader:
                chunk_df.columns = [clean_column_name(c) for c in chunk_df.columns]
                chunk_df = chunk_df.where(pd.notnull(chunk_df), None)
                rows = list(chunk_df.itertuples(index=False, name=None))
                yield list(chunk_df.columns), rows


def stream_excel_chunks(zip_path: str, member_name: str, chunk_size: int):
    """
    Copia el miembro .xlsx del ZIP a un archivo temporal en disco (streaming,
    sin bufferizar todo en RAM) y lo lee con openpyxl en modo read_only,
    entregando las filas en bloques sin cargar el libro completo en memoria.
    """
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    try:
        with zipfile.ZipFile(zip_path, "r") as z, z.open(member_name) as src, open(tmp_path, "wb") as dst:
            shutil.copyfileobj(src, dst, length=1024 * 1024)

        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row is None:
                return
            columns = [clean_column_name(c) for c in header_row]
            n_cols = len(columns)

            batch = []
            for row in rows_iter:
                if not row:
                    continue
                r_len = len(row)
                if r_len < n_cols:
                    row = row + (None,) * (n_cols - r_len)
                elif r_len > n_cols:
                    row = row[:n_cols]

                if not any(row):
                    continue

                batch.append(row)
                if len(batch) >= chunk_size:
                    yield columns, batch
                    batch = []
            if batch:
                yield columns, batch
        finally:
            wb.close()
    finally:
        os.remove(tmp_path)



def stream_legacy_xls_chunks(zip_path: str, member_name: str, chunk_size: int):
    """
    Fallback para archivos .xls antiguos (formato binario no soportado por
    openpyxl). Son archivos legado y normalmente pequeños; se cargan completos
    con pandas, pero igual se entregan en bloques hacia MySQL como el resto.
    """
    logger.warning(f"Archivo .xls legado detectado ('{member_name}'): se carga completo en memoria (sin streaming).")
    with zipfile.ZipFile(zip_path, "r") as z:
        with z.open(member_name) as f:
            df = pd.read_excel(io.BytesIO(f.read()), dtype=str)

    df.columns = [clean_column_name(c) for c in df.columns]
    df = df.where(pd.notnull(df), None)
    columns = list(df.columns)
    rows = list(df.itertuples(index=False, name=None))
    for i in range(0, len(rows), chunk_size):
        yield columns, rows[i:i + chunk_size]


def get_chunks_iterator(zip_path: str, member_name: str, chunk_size: int):
    lower = member_name.lower()
    if lower.endswith(".xlsx"):
        return stream_excel_chunks(zip_path, member_name, chunk_size)
    elif lower.endswith(".xls"):
        return stream_legacy_xls_chunks(zip_path, member_name, chunk_size)
    else:
        return stream_csv_chunks(zip_path, member_name, chunk_size)


def parse_filename_metadata(filename: str) -> tuple:
    """
    Extrae (anio, mes_nombre) a partir del nombre del archivo ZIP.
    Ejemplo: '01_Importaciones_2018_Enero.zip' -> (2018, 'Enero')
    """
    match = re.search(r'(\d{4})_([A-Za-z]+)', filename)
    if match:
        return int(match.group(1)), match.group(2)
    return None, None


def process_zip_files(file_type: str, source_dir: str, db_manager: MySQLManager, limit_files=None, batch_size=5000):
    """
    Procesa todos los archivos ZIP contenidos en el directorio especificado,
    leyendo y cargando cada uno en bloques (streaming) hacia MySQL.
    """
    table_names = {
        "importaciones": "temporal_impo",
        "impo": "temporal_impo",
        "exportaciones": "temporal_expo",
        "expo": "temporal_expo"
    }
    table_name = table_names.get(file_type.lower(), f"temporal_{file_type.lower()}")

    # La limpieza se realiza de forma quirúrgica por archivo (DELETE por archivo_origen)
    # tanto en la Fase 2 (idempotencia antes de insertar) como al finalizar la Fase 4.

    if not os.path.exists(source_dir):
        logger.info(f"Directorio de origen '{source_dir}' no existe. Omitiendo.")
        return

    zip_files = sorted(glob.glob(os.path.join(source_dir, "*.zip")))
    if not zip_files:
        logger.info(f"No hay archivos ZIP pendientes en '{source_dir}'.")
        return

    if limit_files:
        zip_files = zip_files[:limit_files]

    total_zips = len(zip_files)
    logger.info(f"Encontrados {total_zips} archivos ZIP para procesar en '{source_dir}' -> Tabla MySQL: '{table_name}'.")

    # Asegurar índice sobre archivo_origen (usado en el borrado por idempotencia) para tablas ya existentes
    db_manager.ensure_index(table_name)

    for idx, zip_path in enumerate(zip_files, start=1):
        filename = os.path.basename(zip_path)
        logger.info(f"=== [{idx}/{total_zips}] Iniciando procesamiento (streaming): {filename} ===")

        try:
            with zipfile.ZipFile(zip_path, "r") as z:
                member_name = pick_valid_member(z)
            logger.info(f"Leyendo en bloques el archivo interno '{member_name}' de '{filename}'...")

            year, month = parse_filename_metadata(filename)
            chunks_iter = get_chunks_iterator(zip_path, member_name, batch_size)

            rows_inserted = db_manager.load_file_chunks(
                table_name=table_name,
                chunks_iter=chunks_iter,
                anio_proceso=year,
                mes_proceso=month,
                archivo_origen=filename,
            )

            if rows_inserted == 0:
                logger.warning(f"El archivo {filename} no contenía filas de datos.")

            # Ya cargado en MySQL (tabla temporal): el ZIP no se vuelve a necesitar.
            # Se intenta eliminar para liberar espacio; si falla, no se detiene el
            # proceso (el archivo quedará pendiente de limpieza manual más tarde).
            try:
                os.remove(zip_path)
                logger.info(
                    f"¡Éxito! [{idx}/{total_zips}] Archivo {filename} procesado ({rows_inserted} filas) y "
                    f"eliminado tras la carga. Faltan {total_zips - idx}.\n"
                )
            except OSError as e:
                logger.warning(f"Archivo {filename} procesado ({rows_inserted} filas) pero no se pudo eliminar: {e}")

        except Exception as e:
            logger.error(f"Error procesando el archivo '{filename}': {e}", exc_info=True)


def main():
    parser = argparse.ArgumentParser(description="ETL Fase 2: Carga en streaming de ZIPs de la DIAN hacia tablas temporales en MySQL (sin transformar datos).")
    parser.add_argument("--limit-files", type=int, default=None, help="Limita la cantidad de archivos a procesar por carpeta.")
    parser.add_argument("--type", choices=["impo", "expo", "all"], default="all", help="Tipo de archivos a procesar: impo, expo o all.")
    parser.add_argument("--batch-size", type=int, default=5000, help="Tamaño de bloque tanto para lectura (chunksize) como para inserción masiva en MySQL.")
    args = parser.parse_args()

    logger.info("Iniciando Fase 02: Carga en streaming hacia MySQL (localhost)...")
    db_manager = MySQLManager()

    if args.type in ["impo", "all"]:
        logger.info("--- Procesando Importaciones ---")
        process_zip_files(
            file_type="importaciones",
            source_dir=IMPO_DIR,
            db_manager=db_manager,
            limit_files=args.limit_files,
            batch_size=args.batch_size
        )

    if args.type in ["expo", "all"]:
        logger.info("--- Procesando Exportaciones ---")
        process_zip_files(
            file_type="exportaciones",
            source_dir=EXPO_DIR,
            db_manager=db_manager,
            limit_files=args.limit_files,
            batch_size=args.batch_size
        )

    logger.info("=== Fase 02 completada exitosamente ===")


if __name__ == "__main__":
    main()
